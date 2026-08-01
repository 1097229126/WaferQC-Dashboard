"""
业务逻辑层 - 适配现有数据库表结构
"""
from sqlalchemy.orm import Session
from typing import List, Dict, Optional, Tuple
import pandas as pd
import os

from app.models.models import Wafer, Measurement
from app.repositories.repository import WaferRepository, MeasurementRepository
from app.schemas.schemas import WaferWithStats


class WaferService:
    """晶圆业务服务"""
    
    def __init__(self, db: Session):
        self.db = db
        self.wafer_repo = WaferRepository()
        self.measurement_repo = MeasurementRepository()
    
    def get_wafers_with_stats(self, skip: int = 0, limit: int = 100, sort_by: str = None, sort_order: str = None) -> Tuple[List[WaferWithStats], int]:
        """
        获取晶片列表及其统计信息
        
        参数:
            skip: 跳过记录数
            limit: 返回记录数
            sort_by: 排序字段 (wafer_no, conc_mean, conc_max, conc_min, conc_uniformity, conc_tolerance, thick_mean, thick_max, thick_min, thick_uniformity, thick_tolerance)
            sort_order: 排序方向 (asc=正序, desc=倒序)
            
        返回:
            (晶圆列表, 总数)
        """
        wafers, total = self.wafer_repo.get_all_wafers(self.db, skip, limit, sort_by, sort_order)
        
        # 为每个晶圆计算统计数据
        wafer_list = []
        for wafer in wafers:
            stats = self.wafer_repo.get_wafer_with_stats(self.db, wafer.wafer_no)
            if stats:
                wafer_data = WaferWithStats(
                    id=stats["wafer"].id,
                    wafer_no=stats["wafer"].wafer_no,
                    original_grade=stats["wafer"].original_grade,
                    concentration_target=stats["wafer"].concentration_target,
                    thickness_target=stats["wafer"].thickness_target,
                    created_at=stats["wafer"].created_at,
                    updated_at=stats["wafer"].updated_at,
                    avg_concentration=stats["avg_concentration"],
                    avg_thickness=stats["avg_thickness"],
                    measurement_count=stats["measurement_count"],
                    # 浓度统计指标（基于设备1）
                    conc_mean=stats.get("conc_mean"),
                    conc_max=stats.get("conc_max"),
                    conc_min=stats.get("conc_min"),
                    conc_uniformity=stats.get("conc_uniformity"),
                    conc_tolerance=stats.get("conc_tolerance"),
                    # 厚度统计指标（基于设备1）
                    thick_mean=stats.get("thick_mean"),
                    thick_max=stats.get("thick_max"),
                    thick_min=stats.get("thick_min"),
                    thick_uniformity=stats.get("thick_uniformity"),
                    thick_tolerance=stats.get("thick_tolerance")
                )
                wafer_list.append(wafer_data)
        
        return wafer_list, total
    
    def get_wafer_by_no(self, wafer_no: str) -> Optional[Wafer]:
        """
        根据晶片号获取晶片信息
        
        参数:
            wafer_no: 晶片号
            
        返回:
            晶片信息
        """
        return self.wafer_repo.get_wafer_by_no(self.db, wafer_no)
    
    def create_wafer(self, wafer_data: Dict) -> Wafer:
        """
        创建晶片记录
        
        参数:
            wafer_data: 晶片数据字典
            
        返回:
            创建的晶片记录
        """
        return self.wafer_repo.create_wafer(self.db, wafer_data)
    
    def update_wafer(self, wafer_no: str, wafer_data: Dict) -> Wafer:
        """
        更新晶片记录
        
        参数:
            wafer_no: 晶片号
            wafer_data: 晶片数据字典
            
        返回:
            更新后的晶片记录
        """
        return self.wafer_repo.update_wafer(self.db, wafer_no, wafer_data)
    
    def delete_wafer(self, wafer_no: str) -> None:
        """
        删除晶片记录
        
        参数:
            wafer_no: 晶片号
        """
        self.wafer_repo.delete_wafer(self.db, wafer_no)
    
    def create_measurement(self, measurement_data: Dict) -> Measurement:
        """
        创建测量记录
        
        参数:
            measurement_data: 测量数据字典
            
        返回:
            创建的测量记录
        """
        return self.measurement_repo.create_measurement(self.db, measurement_data)
    
    def update_measurement(self, measurement_id: int, measurement_data: Dict) -> Measurement:
        """
        更新测量记录
        
        参数:
            measurement_id: 测量记录ID
            measurement_data: 测量数据字典
            
        返回:
            更新后的测量记录
        """
        return self.measurement_repo.update_measurement(self.db, measurement_id, measurement_data)
    
    def delete_measurement(self, measurement_id: int) -> None:
        """
        删除测量记录
        
        参数:
            measurement_id: 测量记录ID
        """
        self.measurement_repo.delete_measurement(self.db, measurement_id)
    
    def add_measurements_to_wafer(self, wafer_no: str, measurements: List[Dict]) -> Wafer:
        """
        为晶片添加测量数据
        
        参数:
            wafer_no: 晶片号
            measurements: 测量数据列表
            
        返回:
            更新后的晶片记录
        """
        wafer = self.wafer_repo.get_wafer_by_no(self.db, wafer_no)
        if not wafer:
            raise ValueError(f"晶片不存在: {wafer_no}")
        
        # 添加测量数据
        for measurement in measurements:
            measurement['wafer_no'] = wafer_no
        self.measurement_repo.bulk_create_measurements(self.db, measurements)
        
        return wafer
    
    def generate_import_template(self, output_path: str) -> str:
        """
        生成Excel导入模板
        
        参数:
            output_path: 输出文件路径
            
        返回:
            生成的文件路径
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 创建工作簿
            wb = Workbook()
            
            # 创建设备1的Sheet
            ws1 = wb.active
            ws1.title = '数据1'
            
            # 设置表头（双层）
            headers_level1 = ['晶片号', '晶片原等级', '浓度目标', '厚度目标'] + ['浓度'] * 25 + ['厚度'] * 25
            headers_level2 = ['Unnamed: 0_level_1', 'Unnamed: 1_level_1', 'Unnamed: 2_level_1', 'Unnamed: 3_level_1'] + [f'P{i}' for i in range(1, 26)] + [f'T{i}' for i in range(1, 26)]
            
            # 写入第一层表头
            for col_idx, header in enumerate(headers_level1, 1):
                cell = ws1.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # 写入第二层表头
            for col_idx, header in enumerate(headers_level2, 1):
                cell = ws1.cell(row=2, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # 添加示例数据
            example_data = [
                ['TK1', 'D', 1.5e15, 10.0] + [1.5e15 + i*1e13 for i in range(1, 26)] + [10.0 + i*0.1 for i in range(1, 26)],
                ['TK2', 'NG', 1.5e15, 10.0] + [1.5e15 + i*1e13 for i in range(1, 26)] + [10.0 + i*0.1 for i in range(1, 26)],
                ['TK3', 'D', 1.5e15, 10.0] + [1.5e15 + i*1e13 for i in range(1, 26)] + [10.0 + i*0.1 for i in range(1, 26)],
            ]
            
            for row_idx, row_data in enumerate(example_data, 3):
                for col_idx, value in enumerate(row_data, 1):
                    ws1.cell(row=row_idx, column=col_idx, value=value)
            
            # 创建设备2的Sheet
            ws2 = wb.create_sheet(title='数据2')
            
            # 设置表头（双层）- 设备2使用不同的列名
            headers_level1_dev2 = ['Wafer ID', '衬底级别', '浓度目标', '厚度目标'] + ['浓度'] * 25 + ['厚度'] * 25
            
            # 写入第一层表头
            for col_idx, header in enumerate(headers_level1_dev2, 1):
                cell = ws2.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # 写入第二层表头
            for col_idx, header in enumerate(headers_level2, 1):
                cell = ws2.cell(row=2, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # 添加示例数据（设备2的数据略有不同）
            example_data_dev2 = [
                ['TK1', 'D', 1.5e15, 10.0] + [1.5e15 + i*1e13 + 1e12 for i in range(1, 26)] + [10.0 + i*0.1 + 0.05 for i in range(1, 26)],
                ['TK2', 'NG', 1.5e15, 10.0] + [1.5e15 + i*1e13 + 1e12 for i in range(1, 26)] + [10.0 + i*0.1 + 0.05 for i in range(1, 26)],
                ['TK3', 'D', 1.5e15, 10.0] + [1.5e15 + i*1e13 + 1e12 for i in range(1, 26)] + [10.0 + i*0.1 + 0.05 for i in range(1, 26)],
            ]
            
            for row_idx, row_data in enumerate(example_data_dev2, 3):
                for col_idx, value in enumerate(row_data, 1):
                    ws2.cell(row=row_idx, column=col_idx, value=value)
            
            # 保存文件
            wb.save(output_path)
            
            print(f"模板已生成: {output_path}")
            return output_path
            
        except Exception as e:
            raise Exception(f"生成模板失败: {str(e)}")
    
    def import_wafers_from_excel(self, file_path: str) -> Dict:
        """
        从Excel文件导入晶片数据
        
        参数:
            file_path: Excel文件路径
            
        返回:
            导入结果统计信息
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        try:
            # 读取两个Sheet（设备1和设备2）
            print("开始读取Excel文件...")
            df_device1 = pd.read_excel(file_path, sheet_name='数据1', header=[0, 1])
            df_device2 = pd.read_excel(file_path, sheet_name='数据2', header=[0, 1])
            print(f"设备1数据: {len(df_device1)} 行, 设备2数据: {len(df_device2)} 行")
            
            imported_wafers = 0
            imported_measurements = 0
            skipped_wafers = 0
            
            # 处理所有唯一的晶片号
            all_wafer_nos = set()
            
            # 兼容设备1和设备2的列名
            wafer_col_1 = None
            if ('晶片号', 'Unnamed: 0_level_1') in df_device1.columns:
                wafer_col_1 = ('晶片号', 'Unnamed: 0_level_1')
            elif ('Wafer ID', 'Unnamed: 0_level_1') in df_device1.columns:
                wafer_col_1 = ('Wafer ID', 'Unnamed: 0_level_1')
            
            wafer_col_2 = None
            if ('晶片号', 'Unnamed: 0_level_1') in df_device2.columns:
                wafer_col_2 = ('晶片号', 'Unnamed: 0_level_1')
            elif ('Wafer ID', 'Unnamed: 0_level_1') in df_device2.columns:
                wafer_col_2 = ('Wafer ID', 'Unnamed: 0_level_1')
            
            if wafer_col_1:
                print("处理设备1的晶片号...")
                wafer_nos_1 = df_device1[wafer_col_1].dropna().unique()
                print(f"设备1晶片号数量: {len(wafer_nos_1)}")
                all_wafer_nos.update(wafer_nos_1)
            
            if wafer_col_2:
                print("处理设备2的晶片号...")
                wafer_nos_2 = df_device2[wafer_col_2].dropna().unique()
                print(f"设备2晶片号数量: {len(wafer_nos_2)}")
                all_wafer_nos.update(wafer_nos_2)
            
            print(f"总晶片号数量: {len(all_wafer_nos)}")
            
            for wafer_no in all_wafer_nos:
                # 确保晶片号是字符串
                wafer_no = str(wafer_no).strip()
                
                # 检查晶片是否已存在
                existing = self.wafer_repo.get_wafer_by_no(self.db, wafer_no)
                
                # 获取晶片信息（优先从设备1获取）
                wafer_info = None
                for df in [df_device1, df_device2]:
                    try:
                        # 兼容不同设备的列名：晶片号/Wafer ID
                        wafer_col = None
                        if ('晶片号', 'Unnamed: 0_level_1') in df.columns:
                            wafer_col = ('晶片号', 'Unnamed: 0_level_1')
                        elif ('Wafer ID', 'Unnamed: 0_level_1') in df.columns:
                            wafer_col = ('Wafer ID', 'Unnamed: 0_level_1')
                        
                        if wafer_col:
                            row = df[df[wafer_col].astype(str).str.strip() == wafer_no]
                            if not row.empty:
                                wafer_info = row.iloc[0]
                                break
                    except Exception as e:
                        print(f"处理晶片 {wafer_no} 时出错: {e}")
                        continue
                
                if wafer_info is None:
                    continue
                
                # 创建或更新晶片记录
                if not existing:
                    # 提取晶片信息（兼容不同设备的列名）
                    # 等级字段：晶片原等级/衬底级别 是同一个概念
                    original_grade_col = None
                    if ('晶片原等级', 'Unnamed: 1_level_1') in wafer_info.index:
                        original_grade_col = ('晶片原等级', 'Unnamed: 1_level_1')
                    elif ('衬底级别', 'Unnamed: 1_level_1') in wafer_info.index:
                        original_grade_col = ('衬底级别', 'Unnamed: 1_level_1')
                    
                    original_grade = str(wafer_info[original_grade_col]) if original_grade_col and pd.notna(wafer_info[original_grade_col]) else None
                    
                    # 浓度目标和厚度目标
                    concentration_target = float(wafer_info[('浓度目标', 'Unnamed: 2_level_1')]) if pd.notna(wafer_info[('浓度目标', 'Unnamed: 2_level_1')]) else None
                    thickness_target = float(wafer_info[('厚度目标', 'Unnamed: 3_level_1')]) if pd.notna(wafer_info[('厚度目标', 'Unnamed: 3_level_1')]) else None
                    
                    wafer_data = {
                        "wafer_no": wafer_no,
                        "original_grade": original_grade,
                        "concentration_target": concentration_target,
                        "thickness_target": thickness_target
                    }
                    self.wafer_repo.create_wafer(self.db, wafer_data)
                    imported_wafers += 1
                else:
                    skipped_wafers += 1
                
                # 处理设备1的数据
                measurements_device1 = self._extract_measurements_from_row(df_device1, wafer_no, equipment=1)
                if measurements_device1:
                    self.measurement_repo.bulk_create_measurements(self.db, measurements_device1)
                    imported_measurements += len(measurements_device1)
                
                # 处理设备2的数据
                measurements_device2 = self._extract_measurements_from_row(df_device2, wafer_no, equipment=2)
                if measurements_device2:
                    self.measurement_repo.bulk_create_measurements(self.db, measurements_device2)
                    imported_measurements += len(measurements_device2)
            
            self.db.commit()
            
            return {
                "success": True,
                "imported_wafers": imported_wafers,
                "skipped_wafers": skipped_wafers,
                "imported_measurements": imported_measurements,
                "message": f"成功导入 {imported_wafers} 个晶片，{imported_measurements} 条测量数据"
            }
            
        except Exception as e:
            self.db.rollback()
            import traceback
            error_trace = traceback.format_exc()
            print(f"导入错误:\n{error_trace}")
            raise Exception(f"导入失败: {str(e)}")
    
    def _extract_measurements_from_row(self, df: pd.DataFrame, wafer_no: str, equipment: int) -> List[Dict]:
        """
        从DataFrame行中提取测量数据
        
        参数:
            df: DataFrame对象
            wafer_no: 晶片号
            equipment: 设备编号（1或2）
            
        返回:
            测量数据列表
        """
        measurements = []
        
        # 兼容不同设备的列名
        wafer_col = None
        if ('晶片号', 'Unnamed: 0_level_1') in df.columns:
            wafer_col = ('晶片号', 'Unnamed: 0_level_1')
        elif ('Wafer ID', 'Unnamed: 0_level_1') in df.columns:
            wafer_col = ('Wafer ID', 'Unnamed: 0_level_1')
        
        if wafer_col is None:
            print(f"警告：无法找到晶片号列，跳过设备 {equipment}")
            return measurements
        
        # 使用类型转换确保比较正确
        try:
            row = df[df[wafer_col].astype(str).str.strip() == wafer_no]
        except Exception as e:
            print(f"提取晶片 {wafer_no} 的测量数据时出错: {e}")
            return measurements
        
        if row.empty:
            return measurements
        
        row_data = row.iloc[0]
        
        # 提取浓度测量点 P1-P25
        for i in range(1, 26):
            col_name = ('浓度', f'P{i}')
            if col_name in df.columns and pd.notna(row_data[col_name]):
                value = float(row_data[col_name])
                measurements.append({
                    "wafer_no": wafer_no,
                    "measurement_type": 1,  # 浓度
                    "point_number": i,
                    "value": value,
                    "measurement_equipment": equipment
                })
        
        # 提取厚度测量点 T1-T25
        for i in range(1, 26):
            col_name = ('厚度', f'T{i}')
            if col_name in df.columns and pd.notna(row_data[col_name]):
                value = float(row_data[col_name])
                measurements.append({
                    "wafer_no": wafer_no,
                    "measurement_type": 2,  # 厚度
                    "point_number": i,
                    "value": value,
                    "measurement_equipment": equipment
                })
        
        return measurements
